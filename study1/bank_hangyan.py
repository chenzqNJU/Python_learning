import tushare as ts
import numpy as np
import pandas as pd
import os
import gc
os.chdir("e:\\stock")   #修改当前工作目录
from datetime import datetime, timedelta
import time

a=pd.read_excel('guzhi.xlsx',header=None)

t=a.iloc[:,-1]
t=t.str.replace('万','')
t=t.str.replace('亿','')
t=t.astype(float)
a.iloc[:,-1]=t
t=a.iloc[:,-2]
t=t.str.replace('万','')
t=t.str.replace('亿','')
t=t.astype(float)
a.iloc[:,-2]=t

a['6']=a.iloc[:,-1]*2-a.iloc[:,-2]
a['7']=a.iloc[:,-1]*2-a.iloc[:,-2]

t1=a.iloc[:,3]
t=a['6'].astype(str)
t2=t[t1.str.contains('万')].apply(lambda x:x+'万')
t3=t[t1.str.contains('亿')].apply(lambda x:x+'亿')
t4=t2.append(t3)
t4=t4.sort_index()
a['6']=t4

t=a['7'].astype(str)
t2=t[t1.str.contains('万')].apply(lambda x:x+'万')
t3=t[t1.str.contains('亿')].apply(lambda x:x+'亿')
t4=t2.append(t3)
t4=t4.sort_index()
a['7']=t4

a.to_excel('temp11.xlsx')

############################################################
############################################资产负债表
#导入资产负债表、及指标目录
zc = pd.read_csv('zichanfuzhaibiao.csv',encoding='GBK')
menu = pd.read_excel('bank_hangyan.xlsx',sheetnames='Sheet1')
#按code 日期 会计日期排序，由于年报会在未来重复披露多次，保留第一次披露年报
zc.sort_values(['secID','endDate','endDateRep'],inplace=True)
zc.drop_duplicates(subset=['secID','endDate'], keep='first', inplace=True)
zc = zc.drop(zc.columns[0],axis=1)#删除第一个变量
zc.set_index(['secID','endDate'],inplace=True)
zc.endDateRep=pd.to_datetime(zc.endDateRep, format='%Y-%m-%d')
#提取zc的变量名，构造表
name=zc.columns#index格式
name.names=['name']#命名
name=name.to_frame()
#将变量名 与变量的含义连接
menu1=pd.merge(name,menu,left_on='name',right_on='名称',how='left').iloc[:,[0,3]]
menu2=menu1.set_index('name').unstack().unstack()
zc1=menu2.append(zc)

############################################利润表
lr = pd.read_csv('lirunbiao.csv',encoding='GBK')
menu = pd.read_excel('bank_hangyan.xlsx','Sheet2')

lr.sort_values(['secID','endDate','endDateRep'],inplace=True)
lr.drop_duplicates(subset=['secID','endDate'], keep='first', inplace=True)
lr = lr.drop(lr.columns[0],axis=1)
lr.set_index(['secID','endDate'],inplace=True)
lr.endDateRep=pd.to_datetime(lr.endDateRep, format='%Y-%m-%d')

name=lr.columns
name.names=['name']
name=name.to_frame()
menu1_=pd.merge(name,menu,left_on='name',right_on='名称',how='left').iloc[:,[0,3]]
menu1_=menu1_.merge(menu1,on='name',how='left')
#将利润表变量的重复变量 用资产表里面的替换（有些变量说明很长）
menu1_.loc[~(menu1_.描述_y.isnull()),'描述_x']=menu1_.loc[~(menu1_.描述_y.isnull())].描述_y
menu1=menu1_.drop('描述_y',axis=1).rename(columns={'描述_x':'描述'})
menu2=menu1.set_index('name').unstack().unstack()
lr1=menu2.append(lr)
############################################现金流量表
xj = pd.read_csv('xianjinliuliangbiao.csv',encoding='GBK')
menu = pd.read_excel('bank_hangyan.xlsx','Sheet3')

xj.sort_values(['secID','endDate','endDateRep'],inplace=True)
xj.drop_duplicates(subset=['secID','endDate'], keep='first', inplace=True)
xj = xj.drop(xj.columns[0],axis=1)
xj.set_index(['secID','endDate'],inplace=True)
xj.endDateRep=pd.to_datetime(xj.endDateRep, format='%Y-%m-%d')

name=xj.columns
name.name='name'
name=name.to_frame()
menu1_=pd.merge(name,menu,left_on='name',right_on='名称',how='left').iloc[:,[0,3]]
menu1_=menu1_.merge(menu1,on='name',how='left')
menu1_.loc[~(menu1_.描述_y.isnull()),'描述_x']=menu1_.loc[~(menu1_.描述_y.isnull())].描述_y
menu1=menu1_.drop('描述_y',axis=1).rename(columns={'描述_x':'描述'})
menu2=menu1.set_index('name').unstack().unstack()
xj1=xj.append(xj)

############################################################################
######################################导出3张报表
#变量名换成中文描述
xj.columns=xj1.loc['描述']
#转化形态，index为指标+时间，列为银行
xj2=xj.unstack(level=[0,1]).unstack(level=1)
#读取各银行的简称
jiancheng=xj2.loc[[('证券简称','2017-09-30')],:]
#删除前面没用的指标，由于是2级index，对第一层进行查找
t=xj2.index.levels[0]
n=t.tolist().index('货币代码')
xj2=xj2.drop(t[:n+1])
xj2=jiancheng.append(xj2)

zc.columns=zc1.loc['描述']
#这里存在列名一样的，进行删除
zc=zc.loc[:,~zc.columns.duplicated(False)]
zc2=zc.unstack(level=[0,1]).unstack(level=1)
jiancheng=xj2.loc[[('证券简称','2017-09-30')],:]
t=zc2.index.levels[0]
n=t.tolist().index('货币代码')
zc2=zc2.drop(t[:n+1])
zc2=jiancheng.append(zc2)

lr.columns=lr1.loc['描述']
lr2=lr.unstack(level=[0,1]).unstack(level=1)
jiancheng=xj2.loc[[('证券简称','2017-09-30')],:]
t=lr2.index.levels[0]
n=t.tolist().index('货币代码')
lr2=lr2.drop(t[:n+1])
lr2=jiancheng.append(lr2)

##########################################批量导出3张报表
writer = pd.ExcelWriter('bank_hangyan\\book.xlsx')
zc2.to_excel(writer, '资产负债表')
lr2.to_excel(writer, '利润表')
xj2.to_excel(writer, '现金流量表')
writer.save()

##########################################农商行报表
## 目录下所有文件
def all_path(dirname):
    result = []
    for maindir, subdir, file_name_list in os.walk(dirname):
        for filename in file_name_list:
            apath = os.path.join(maindir, filename)
            result.append(apath)
    return result

t='E:\stock\五家上市农商行2014-2017所有报表\行业改革发展调研基层法人数据表_61无锡农商行'
t='\\'.join(t.split('\\'))
result=all_path(os.path.dirname(t))

z=[]
def find(str):
    z = []
    for each in result:
        if str in each and '.xml' not in each:
            z.append(each)
    return z
length=len(z)


z=[]
for each in result:
    if 'S44' in each:
        z.append(each)
length=len(z)

z=[]
for each in result:
    if 'G11_1' in each  or 'G1101' in each:
        z.append(each)


def get_year(z):
    year=range(2014,2018)
    z8=[]
    for n,x in enumerate(z):
        temp=[x.count(str(i)) for i in year]
        if temp[1]>0 and temp[2]>0:
            print("error")
        temp[1]*=100;temp[2]*=100;
        z8+=[(year[np.argmax(temp)])]
    return z8
import re
def get_num(z):
    p1 = "(?<=数据表).*?(?=农商)"
    pattern1 = re.compile(p1)  # 我们在编译这段正则表达式
    z4 = [re.sub("\d", "", re.search(pattern1, key).group(0)) for key in z]
    z4 = [key.replace('-','').replace('_','')+'农商行' for key in z4]
    return z4
def get_month(z3):
    month=[]
    z3=[x.replace('1104','') for x in z]
    pattern1 = re.compile("(?<=201[4-7])[0-9]{1,2}")
    pattern2 = re.compile('(?<=年)[\d]{1,2}(?=月)')
    for i in z3:
        i=i.replace('.','')
        if re.search(pattern1,i):
            month+=[re.search(pattern1,i).group(0)]
        elif re.search(pattern2, i):
            month += [re.search(pattern2, i).group(0)]
        else:month+=['12']
    return [int(x) for x in month]

############杠杆率 g44
z=find('G44')[-8:]
final=Final(z)

final=final.dropna()
t=final['Unnamed: 1']
#正则匹配 1.非数字
final1=final[t.str.contains(r'\d\.\D')]
a=pd.pivot_table(final1,values='Unnamed: 2',columns='Unnamed: 1',index=['name','year'])
z1=final1.set_index(['name','year','Unnamed: 1'])
a=a['Unnamed: 2']
a1=a.unstack(level=-1)
add_sheet(final1,'temp','3')

############资本充足率 g40
z=find('G40')[-8:]
final=Final(z)
final=final.dropna()
t=final['Unnamed: 1']
#正则匹配 1.非数字
final1=final[t.str.contains(r'\d\.\D')]
a=pd.pivot_table(final1,values='Unnamed: 2',columns=final1.columns[0],index=['name','year'])
a=final1.set_index(['name','year',final1.columns[0]])
a=a.drop('month',axis=1)

a1=a.unstack(level=1)
add_sheet(a1,'temp','4')

############利率风险 G33
z=find('G33')
z=list(filter(lambda x:'3302' in x and '人民币' in x,z))[-4:]
final=Final(z)
t=final['Unnamed: 1']
#正则匹配 1.非数字
t=t.fillna('.')
final1=final[t.str.contains(r'\d\.\D')]
a=final1.set_index(['name','year',final1.columns[0]])
a=a.iloc[:,:2]
a1=a.unstack(level=1)
add_sheet(a1,'temp','5')
add_sheet(final1,'temp','6')

############流动性风险 G25
z=find('G25')[-3:]
final=Final(z)

t=final.iloc[:,0].values
final=final[[isinstance(x,int) and x > 83 for x in t]]
add_sheet(final,'temp','7')

############流动性期限风险 G21
z=find('G21')[-3:]
final=Final(z)

t=final.iloc[:,1]
t.str.contains('资产合计')
final=final[[isinstance(x,int) and x > 83 for x in t]]
add_sheet(final,'temp','7')

############资产减值 准备 G03
z=find('G03')
final=Final(z)
final=final.fillna('.')
final1=final[final.iloc[:,1].str.contains('贷款损失')]   #合计
final1=final1.drop(final.columns[3:8],axis=1)
final2=final1[final1.year==2014]
final2['Unnamed: 8']=final2['Unnamed: 2'];final2.year=2013
final1=final1.append(final2)

final1=final1[['year','name','Unnamed: 8']]

final1['Unnamed: 8'] = final1['Unnamed: 8'] .astype(float)
a=pd.pivot_table(final1,columns='name',values='Unnamed: 8',index='year')

add_sheet(a,'temp','9')
############存贷款长度分布  G01
z=find('G01')
z=list(filter(lambda x:'G0104' in x or 'G01_4' in x or 'Ⅳ' in x,z))[-4:]
final=Final(z)
final.iloc[:,1]=final.iloc[:,1].fillna('.')
final1=final[final['Unnamed: 1'].str.contains(r'\d\.')]
final1['Unnamed: 1']=final1['Unnamed: 1'].str.replace(r'\d*\.','')

t1=final1[final1.iloc[:,1].str.contains('合计')]
a=pivot(t1,[2,3,4])
t2=final1[final1.year==2017]
add_sheet(a,'temp','10');add_sheet(t2,'temp','11')

############贷款行业分类  G0107
z=find('G01')
z=list(filter(lambda x:'G0107' in x or 'G01_7' in x or 'G017' in x or 'VII部分' in x,z))
final=Final(z)
final.iloc[:,1]=final.iloc[:,1].fillna('.')
final1=final[final['Unnamed: 1'].str.contains(r'\d\.')]
t=final[final['Unnamed: 1'].str.contains(r'2.21\.')]
t=t[t.name=='张家港农商行']
t=pivot(t,[2],col='Unnamed: 1')
add_sheet(t,'temp','12')

t1=final[final['Unnamed: 1'].str.contains(r'2.1') | final['Unnamed: 1'].str.contains(r'2.\d\D')]
t1=t1[t1.year==2017]
t1['Unnamed: 1']=t1['Unnamed: 1'].str.replace(r'\W','')
t1['Unnamed: 1']=t1['Unnamed: 1'].str.replace(r'\d','')
t1.iloc[:,2] = t1.iloc[:,2].astype(float)
a = pd.pivot_table(t1, columns='name', values=final1.columns[2], index='Unnamed: 1')
add_sheet(a,'temp','13')

############贷款质量五分类  G0102
z=find('G01')
z=list(filter(lambda x:'G0102' in x or 'G01_2' in x or 'G012' in x or 'Ⅱ部分' in x,z))
final=Final(z)
final1=final[final['Unnamed: 1'].str.contains('类')]

t1=final1[final1.year==2017]
t1=pivot(t1,[2],ind='Unnamed: 1')

t2=final1[final1.name=='张家港农商行']
t2['Unnamed: 1']=t2['Unnamed: 1'].str.replace(r'\W','')
t2['Unnamed: 1']=t2['Unnamed: 1'].str.replace(r'\d','')
t2=pivot(t2,[2],col='Unnamed: 1')
add_sheet([t1,t2],'temp','14',flag=2)

############支农情况  S45
z=find('S45')[-4:]
final=Final(z)
final.iloc[:, 2] = final.iloc[:, 2].fillna('.')
final1=final[final['Unnamed: 2'].str.contains('1\.1\.')]
final1=final1[(final1.iloc[:,0] ==3) | (final1.iloc[:,0] ==4)]
t2=pivot(final1,[3],col='Unnamed: 2')

final1=final[final['Unnamed: 2'].str.contains('1\.2\.\d[^\.]')]
t3=pivot(final1,[3],col='Unnamed: 2')

final1=final[final['Unnamed: 2'].str.contains('(农村区域)|(银行卡)')]
t4=pivot(final1,[3],col='Unnamed: 2')

final1=final[final['Unnamed: 2'].str.contains('辖内')]
t5=pivot(final1,[3],col='Unnamed: 2')
add_sheet([t2,t3,t4,t5],'temp','15',flag=2)

############资产表  G0100
z=find('G0100')[-5:]
z.pop(1)
final=Final(z)

final1=final[final['Unnamed: 1'].str.contains('资产总计') |
             final['Unnamed: 1'].str.contains('49\.') |
             final['Unnamed: 1'].str.contains('59\.') |
             final['Unnamed: 1'].str.contains('6\d\.')]
t=pivot(final1,[4],col='Unnamed: 1')
add_sheet(t,'temp','16')
################################################
################################################
################################################
t.str.replace(r'\d\.\d*','')

import myfunc
myfunc.search('正则',dirFlag=1)

def pivot(final1,x,col='name',ind='year'):
    for i in x:
        final1.iloc[:,i] = final1.iloc[:,i].astype(float)
    a = pd.pivot_table(final1, columns=col, values=final1.columns[x], index=ind)
    return a



def Final(z):
    df = pd.DataFrame({'num': get_num(z), 'year': get_year(z), 'month': get_month(z)})
    final = pd.DataFrame()
    length=len(z)
    i = 1;temp = pd.read_excel(z[i])
    t = temp.isnull().sum(axis=0)
    t = ((t - t.min()) < 20).sum()
    t=temp.shape[1]
    for i in range(length):
        try:
            data= pd.read_excel(z[i])
        except:
            continue
        data1=data.iloc[:,:t]   #4
        data1.loc[:,'name']=df.num[i]
        data1.loc[:,'year']=df.year[i]
        data1.loc[:,'month'] = df.month[i]
        final=final.append(data1,ignore_index=True)
    final.iloc[:, 1] = final.iloc[:, 1].fillna('.')
    return final

def add_sheet(var,wookbookname,sheetname,flag=1):
    from openpyxl import load_workbook
    book = load_workbook(wookbookname+'.xlsx')
    writer = pd.ExcelWriter(wookbookname+'.xlsx', engine='openpyxl')
    writer.book = book
    if flag==1:var.to_excel(writer,sheetname)
    if flag==2:
        for i in range(len(var)):
            var[i].to_excel(writer, sheetname, startrow=0, startcol=i*15)
    writer.save()

add_sheet(final1,'temp','1')
#############################################s67
t=final.iloc[:,1]
t1=t.str.count('\.')
t2=t.str.count('\d')
t3=(t1<3) & (t2>0)
final1=final[t3]
final1=final1[final1.iloc[:,1].str.startswith('1')]
final1.to_excel('temp.xlsx')

a=final1[final1.iloc[:,1].str.contains('房地产贷款合计')]
add_sheet(a,'temp','1')

a=pd.pivot_table(final1,values='Unnamed: 2',columns='Unnamed: 1',index=['name','year'])
a=final1.set_index(['name','year','Unnamed: 1'])
a=a['Unnamed: 2']
a1=a.unstack(level=-1)
add_sheet(a1,'temp','2')


final1['Unnamed: 2'][637]

t=final['Unnamed: 9']
t[t.str.contains('损失')==True]

t=final.iloc[:,0]
t1=t.apply(lambda x: isinstance(x,int))
final1=final[t1]
final1.rename(columns={final1.columns[0]:'a'}, inplace = True)

final1.columns='a b 各项贷款 正常贷款 正常类 关注类 不良贷款 次级类 可疑类 ' \
               '损失类 name year month'.split(' ')
final1=final1.fillna(0)
final2=pd.pivot_table(final1,index='b',values=final1.columns[2:10],
               columns=['name',"year"],fill_value=0)
final3=final2.swaplevel(i=0,j=1,axis=1)
final3=final3.sort_index(axis=1)
final3.to_excel('e:\\stock\\五家上市农商行2014-2017所有报表\\G1101贷款质量表.xlsx')

import myfunc
myfunc.search('swa')
#判断是否有序号 和 名称 重复
t2=[]
for x in list(set(final1.a)):
    t1=set(final1[final1.a==x].iloc[:,1])
    if len(t1)!=1:print(x)
    t2+=[t1]

t=final1[final1.year==2014]
t.rename(columns={'Unnamed: 2':'Unnamed: 3','Unnamed: 3':'x'},inplace=True)
t.year=2013
final2=t.append(final1)

final2.set_index(['name','year','Unnamed: 1'],inplace=True)
final2=final2.pop('Unnamed: 3')
final3=final2.unstack(level=[0,1])
final3=final3.sort_index(level=0,axis=1)

final3.to_excel('e:\\stock\\五家上市农商行2014-2017所有报表\\S44损益表.xlsx')



import csv
csv_reader = csv.reader(open(t))
for row in csv_reader:
    print(row)

t='E:\\stock\\五家上市农商行2014-2017所有报表\\行业改革发展调研基层' \
  '法人数据表_60常熟农商行\\2014-2017年度1104报表\\2014年12月\\G01资产负债项目统计表附注第Ⅳ部分：存贷款明细报表（二）.xls'
t='E:\\stock\\五家上市农商行2014-2017所有报表\\行业改革发展调研基层' \
  '法人数据表_60常熟农商行\\2014-2017年度1104报表\\2014年12月\\G01资产负债项目统计表附注第Ⅱ部分：贷款质量五级分类情况简表.xls'


f = open(t, 'rb')
lines = f.readlines()
line=lines[2]
for line in lines:
    line = line.decode('gb2312').encode('utf8')
    print(line)
line.decode('utf')

myfunc.search('with')
a=b'3\x9dj\xd4u\x1a\x10\xd8\xfd?y?0,,,,\r\n'
a=b'F\xc7\xf2?\xcd\xc2?\x18\xfa\xe1V\xe4\xdf-?/\x9d\x9ePig?o=\x96\x985\x99\xd3\xd9\xe6\x8a\xf4\x81R\xc6\x88C\xb3[?\xcaU\xcc\xbe\x0cY}\xc1\xf9\x90\xbc?Jj\x7f\x8f\xd41\\\x82\xad,\xaa\xaa\x90\xe4#\xb2\x89\xfal\xc6\xa9{X\xb0\x8e\x05BE?\xceF\xf9\xb3,,,\r\n'

a.decode('utf8','ignore')
a.decode('GBK')

xls_file = xlrd.open_workbook(t)
xls_sheet = xls_file.sheets()[1-1]
row_value = xls_sheet.row_values(2-1)





import dateutil.parser
z=[dateutil.parser.parse(x[1]).month for x in zc.index]

zc['month1']=np.array(z)
zc[zc.month1>3]
zc[np.array(z)==3]


z=zc.columns.name

help(pd.Series.to_frame)
#对12.31有5份 取靠前的一份
a=zc.groupby(zc.index)
b=a['endDateRep'].min()
b=b.to_frame()
c=pd.merge(b,zc,how='left')

c=pd.concat([zc, b],keys=['endDateRep','endDateRep'],axis=1)



t1=zc.iloc[:10,:4]
t2=b.iloc[:10]
c=pd.concat([t2, t1],keys=['endDateRep'],axis=1)
c=pd.concat([t2, t1],axis=1)


c=pd.merge(t2[2:],t1,on='endDateRep',how='left')

c=pd.merge(t2[2:],t1,on='endDateRep',left_index = True,right_index = True,how='left')

result = pd.concat([t1, t2], axis=1,keys=['endDateRep','endDateRep'])




help(pd.DataFrame.to_excel)

help(pd.concat)
import myfunc
myfunc.search('unstack',1)

zc[3:4].index.is_unique

zc.index.values

label=zc.index.levels[1][2]

zc.index.get_loc(zc.index.levels[1][2])
zc.iloc[zc.index.get_loc(label)]

zc.loc[[zc.index.get_level_values(0)[-1]]]

label=zc.index.labels[1]