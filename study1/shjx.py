import tushare as ts
import numpy as np
import pandas as pd
import os
import gc
import myfunc
np.nan<1

os.chdir("e:\\stock")  # 修改当前工作目录
from datetime import datetime, timedelta
import time
help(pd.DataFrame.isnull)
help(pd.DataFrame.count)

myfunc.search('corr')


b'\xe6\xb1\xbd\xe8\xbd\xa6'.decode('utf8')
1319-1491
myfunc.search('eval')
a='search'
b='myfunc'
eval(b).search('s')
a=['a','f']
a.append(['d','d'])
list(['d','f'])
list('f')
a['g']=[3,3]

a=1
# 保存
dict_name = {1: {1: 2, 3: 4}, 2: {3: 4, 4: 5}}
f = open('temp.txt', 'w')
f.write(str(dict_name))
f.close()

# 读取
f = open('temp.txt', 'r')
a = f.read()
dict_name = eval(a)
f.close()

f = open('e:/temp/dw.txt','r')
a = f.read()
a[-2:]
b=eval(a)
len(a)
a=a.replace('benchmark_cumulative_values','ben')
a.find('cumulative_values')
b=a[7687:]
b.find('2010-01-04')

b=b.replace('\n','    ')
c=b.split('    ')
'1.2'.isdigit()

str_1 = "12.3"
str_2 = "Abc"
str_3 = "123Abc"

#用isdigit函数判断是否数字
print(str_1.isdigit())

help(xlrd.open_workbook)

import re
def is_number(num):
  pattern = re.compile(r'^[-+]?[-0-9]\d*\.\d*|[-+]?\.?[0-9]\d*$')
  result = pattern.match(num)
  if result:
    return True
  else:
    return False
d=[x for x in c if is_number(x)]


myfunc.search('re.',1)









temp = pd.read_excel('E:\\stock\\公司公告.xls', sheet_name='公司公告')

import openpyxl
wb= openpyxl.load_workbook('E:\\stock\\公司公告.xlsx',data_only=True)
# 获得所有sheet的名称
print(wb.get_sheet_names())
# 根据sheet名字获得sheet
a_sheet = wb.get_sheet_by_name('公司公告')

# 获得sheet名
print(a_sheet.title)
# 获得当前正在显示的sheet, 也可以用wb.get_active_sheet()
sheet = wb.active

# 获取某个单元格的值，观察excel发现也是先字母再数字的顺序，即先列再行
b4 = sheet['C4']
# 分别返回
print(f'({b4.column}, {b4.row}) is {b4.value}')  # 返回的数字就是int型
# 除了用下标的方式获得，还可以用cell函数, 换成数字，这个表示B4
b4_too = sheet.cell(row=4, column=3)
print(b4_too.value)

help(xlrd.hyperlink_map)
import xlrd
mainData_book = xlrd.open_workbook("E:\\stock\\公司公告.xls", formatting_info=True)
mainData_book = xlrd.open_workbook("E:\\stock\\公司公告.xls")

mainData_sheet = mainData_book.sheet_by_index(0)

for row in range(1, 101):
    row =3
    rowValues = mainData_sheet.row_values(row, start_colx=0, end_colx=8)
    url = rowValues[3]
    url = url.split('"')[1]
import urllib.request

res = urllib.request.urlopen(url)
html = res.read().decode('utf-8')



for row in range(1, 101):
    rowValues = mainData_sheet.row_values(row, start_colx=0, end_colx=8)
    company_name = rowValues[0]
    link = mainData_sheet.hyperlink_map.get((row,3))
    url = '(No URL)' if link is None else link.url_or_path
    print(company_name.ljust(20) + ': ' + url)





for x  in set(['2018-08-10', '2018-07-31', '2018-08-14', '2018-08-09', '2018-08-08', '2018-08-07', '2018-08-06', '2018-08-01', '2018-08-13', '2018-08-03', '2018-08-02']):
    print(x)

df = pd.DataFrame({'key1' : ['a', 'a', 'b', 'b', 'a'],
    'key2' : ['one', 'two', 'one', 'two', 'one'],
    'data1' : np.random.randn(5),
    'data2' : np.random.randn(5)})
df.groupby('key1')['data1'].pct_change()

a={'c':[1],'f':[1,2]}
eval('myfunc.'+a)('s')

myfunc.eval(a)('symbol')


def chinese(text):
    print('jieba分词')


def chinese(text):
    print('jieba分词')

def english(text):
    print('nltk处理对文本进行分词')

def textprocess(file, language):
    text = open(file).read()
    eval(language)(text)

file =
'data.txt'

def english(text):
    print('nltk处理对文本进行分词')


str_func_pair = {'chinese': chinese,
                 'english': english}


def textprocess(file, language):
    text = open(file).read()
    # 字典调用出函数名
    str_func_pair[language](text)

file =
'data.txt'

language =
'english'

textprocess(file, language)


def chinese(text):
    print('jieba分词')


def english(text):
    print('nltk处理对文本进行分词')


def textprocess(file, language):
    text = open(file).read()

    # 字符串调用chinese()或english()函数。
    language(text)


def chinese(text):
    print('jieba分词')


def english(text):
    print('nltk处理对文本进行分词')


str_func_pair = {'chinese': chinese,
                 'english': english}


def textprocess(file, language):
    text = open(file).read()
    # 字典调用出函数名
    str_func_pair[language](text)

# 调用函数
file =
'
data.txt
'
language =
'english'
textprocess(file, language)

help(pd.DataFrame.drop_duplicates)

kongzhiren=pd.read_csv('kongzhiren.csv')
kongzhiren=kongzhiren.iloc[:,1:]
kongzhiren.ticker=kongzhiren.ticker.astype(str)
kongzhiren.ticker=kongzhiren.ticker.str.zfill(6)
kongzhiren=kongzhiren[kongzhiren.concurrentpost>0]
kongzhiren.concurrentpost=kongzhiren.concurrentpost.astype('int')
kongzhiren=kongzhiren.sort_values(['ticker','changeDate'])

kongzhiren.to_csv('kongzhiren.csv',index=None)



a = 1;
b = a + 1;
c = 4 - b


import pdb
a = "aaa"
pdb.set_trace()
b = "bbb"
c = "ccc"
d = a + b + c
print(d)

import numpy as np
class Network:
    def __init__(self,sizes):
        self.num_layers=len(sizes)
        self.sizes=sizes
        self.biases=[np.random.randn(y,1) for y in sizes[1:]]
        self.weights=[np.random.randn(x,y) \
                    for x,y in zip(sizes[1:],sizes[:-1])]
        self.null=[]
sizes=[2,3,1]
list(zip(sizes[1:],sizes[:-1]))

net=Network([2,3,1])
def sigmoid(z):
    return 1.0/(1.0+np.exp(-z))

def foo(s):
    n = int(s)
    assert n != 0, 'n is zero!'
    return 10 / n

def main():
    foo('0')

main()

import logging
logging.basicConfig(level=logging.INFO)
s = '0'
n = int(s)
pdb.set_trace()
a=1
b=2
print(10 / n)

np.sum()

if a is None:

import myfunc
from myfunc import *

delete()
search_('dff')
add_(2, 3)
dir()
myfunc.delete()
myfunc.search('groupby')

# python 不支持重载 即 import pandas 两次，则后一次不运行在默认情况下，只是在每次会话的第一次运行。在第一次导入之后，其他的导入都不会再工作，
# 甚至在另一个窗口中改变并保存了模块的源代码文件也不行。
import myfunc
from importlib import reload

reload(myfunc)

import sys

sys.path.append(r'C:\\Users\\chenzq\\Documents\\9_23\\subfunction')
from funcSimple import *
from funcSimple_ import *
import funcSimple

funcSimple.add_(2, 3)
x = 1
y = 2
a = add(x, y)
print(a)

sys.path.append(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from datetime import datetime

a = datetime.strptime('2013-01', '%Y-%m')
str(a)[:10]
a.strftime('%F')


# 删除变量
def de():
    for i in dir():
        if not i.startswith('__'):
            locals().pop(i)


de()

a = 1;
b = 2;
c = 3


def __clear_env():
    for key in globals().keys():
        if not key.startswith("__"):  # 排除系统内建函数
            globals().pop(key)


__clear_env()

z = list(globals().keys())
for key in globals().keys():
    if not key.startswith("__"):
        globals().pop(key)

dates = ['2017-06-20', '2017-06-21', \
         '2017-06-22', '2017-06-23', '2017-06-24', '2017-06-25', '2017-07-26', '2017-07-27']

ts = pd.Series(np.random.randn(8), index=pd.to_datetime(dates))
ts['2017-07']
ts.name = 'df1'
a = ts.to_frame()
a.index.name = 'date'
a['date'] = a.index
a['df2'] = np.random.randn(8)
#################根据月份进行分类：6月、7月
b = a.groupby(pd.Grouper(freq='M'))
#################分别对2个月计算：date的最大值、df1的多个值
b.agg({'date': np.max, 'df1': ['min', 'max', 'mean', 'std']})
###################对df1 df2计算a b两个特质
ftuples = [('a', 'mean'), ('b', np.var)]  # (名称，函数)
b['df1', 'df2'].agg(ftuples)
###################对df1 df2计算3个特质
functions = ['count', 'mean', 'max']
b['df1', 'df2'].agg(functions)
###################分类下，取最后一个 第一个，最大值, 非空值个数
b.first()
b.max()
b.last()
b.count()

df = pd.DataFrame({'key1': list('aaaab'),
                   'key2': ['one', 'two', 'one', 'two', 'one'],
                   'data1': np.random.randn(5),
                   'data2': np.random.randn(5)})

c = df.groupby('key1')

df.groupby('key1', as_index=False).agg('sum')
df.groupby('key1')[['data1']].agg('sum')

#######################################月末日期序列  MS是月初
pd.date_range('2011-01-03', periods=5, freq='M')

pd.Timestamp('2014-01-02') + MonthEnd(n=1)

#########################################时间偏移
d = datetime(2008, 8, 18, 9, 0)
from dateutil.relativedelta import relativedelta

t = d + relativedelta(months=4, days=5)

from pandas.tseries.offsets import *

d + DateOffset(months=4, days=5)
pd.Timestamp('2008-12-23 09:00:00') + MonthEnd(-1)
pd.Timestamp('2014-01-02') + MonthEnd(n=2)

###############################dateutil模块 时间判别
import dateutil.parser
import dateutil.rrule
from dateutil.parser import parse
from dateutil.rrule import *

dateutil.parser.parse('19/May/2017').date()


def getDateTime(s):
    d = dateutil.parser.parse(s)
    return d


list(dateutil.rrule.rrule(freq='MONTHLY', dtstart=dateutil.parser.parse('2013-05-19'),
                          until=dateutil.parser.parse('2013-08-20')))
list(rrule(MONTHLY, bymonth=1, dtstart=parse('2013-05-19'), until=parse('2013-08-20')))
list(rrule(MONTHLY, dtstart=parse('2013-05-31'), until=parse('2013-08-20')))

help(rrule)
from dateutil.relativedelta import relativedelta

d = datetime.now()

import dateutil.DateUtil

dateutil.toString.toString(d)
# Next Friday
print(d + relativedelta(weekday=FR))
# Last Friday
print(d + relativedelta(weekday=FR(-1)))

import calendar

calendar.FRIDAY
# 每年的第1 100天
list(rrule(YEARLY, count=4, interval=3, byyearday=(1, 100, 200), dtstart=parse("19970101T090000")))
# 每个月第一个 最后一个星期
list(rrule(MONTHLY, interval=2, count=10, byweekday=(SU(1), SU(-1)), dtstart=parse("19970907T090000")))
# 月份最后一天
list(rrule(MONTHLY, interval=2, count=10, bymonthday=(-1), dtstart=parse("20180324T090000")))

######################################################
######################################################
######################################################
temp = pd.read_excel('E:\\stock\\风险承担\\图表.xlsx', sheetname='Sheet4')
temp=temp[1:]
try:
    temp = temp[temp.标准 != '.']
except:
    temp


def star(var):
    b = []
    for x in temp.iloc[:, var]:
        if isinstance(x, float):
            x1 = str(x)
            if x < 0.1:
                x1 = str(x) + ' *'
            if x < 0.05:
                x1 = str(x) + ' **'
            if x < 0.01:
                x1 = str(x) + ' ***'
        else:
            x1 = str(x) + ' ***'
        b.append(x1)
    temp.iloc[:, var] = b
star(5)
temp = temp[['参数','估计','t 值','Pr > |t|']]
temp1=temp
add_sheet('E:\\stock\\风险承担\\图表','sheet5')

def add_sheet(wookbookname, sheetname):
    from openpyxl import load_workbook
    book = load_workbook(wookbookname + '.xlsx')
    writer = pd.ExcelWriter(wookbookname + '.xlsx', engine='openpyxl')
    writer.book = book
    temp1.to_excel(writer, sheetname)
    writer.save()


a = len(temp.columns) - 1;
b = int(len(temp.columns) / 2 - 1)
star(a);
star(b)

temp1 = temp.iloc[:, [0, 1, 3, 4, 6, 8, 9]]
temp1.set_index('参数', inplace=True)
a = [['ROA_std'] * 3 + ['Return_std'] * 3] + [['估计', 't值', 'p值'] * 2]
temp1.columns = pd.MultiIndex.from_tuples(list(zip(*a)))

temp1.to_excel('temp1.xlsx', '会计变量')
add_sheet('temp1', '环境变量')
add_sheet('temp1', '股东变量')
add_sheet('temp1', '会计变量')
add_sheet('temp1', '哑变量')
add_sheet('temp1', '所有变量')
##################################导入到两个sheet工作簿
writer = pd.ExcelWriter('temp1.xlsx')
temp1.to_excel(writer, 'sheet1')
temp1.to_excel(writer, 'sheet2')
writer.save()
##################################

import os


def file_name(key, file_dir='C:\\Users\\chenzq\\Documents\\9_23'):
    temp = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] in ['.py', '.txt']:
                temp.append(os.path.join(root, file))
    for dir in temp:
        s = 0
        # dir=file[8]
        try:
            with open(dir, encoding='utf-8') as f:
                for i in f:
                    if key in i:
                        print(i), print(dir)
                        s += 1
                # if s == 0:
                # print("don't match it！")
        except:
            try:
                with open(dir, encoding='gbk') as f:
                    for i in f:
                        if key in i:
                            print(i),
                            s += 1
                    # if s == 0:
                    # print("don't match it！")
            except:
                print(dir)


file_name('jindu')

add_sheet('temp1', 'dd')

ts[ts.index > '2017-06-22']
ts.index.is_unique
set(ts.index)
ts.index[1].date()
a = pd.Timestamp('2017-06-21 00:00:00')
a = datetime(2017, 6, 21)
ts.index[1] == a
a.time()

temp.iloc[:, 4]
temp.估计.mean(skipna=True)
np.array(temp.估计).nanmean()
np.nanmean(np.array(temp.估计))

temp = temp.dropna()
a = temp['P值']
type(a[1])
isinstance(a[1], str)
a[1].replace('<', '0')
b = []
x = a[2]
for x in a:
    if isinstance(x, float):
        if x < 0.1:
            x1 = str(x) + ' *'
        if x < 0.05:
            x1 = str(x) + ' **'
        if x < 0.01:
            x1 = str(x) + ' ***'
    else:
        x1 = str(x) + ' ***'
    b.append(x1)
a.P值.apply(lambda x: x[:3])
temp['P值'] = b

temp.columns
a = temp.drop('t 值', axis=1)
a = a.set_index('参数')
b = a.stack()
c = pd.DataFrame({'dfdf': b})
c.to_excel('tem.xlsx')

df = ts.get_today_ticks('000852')
df1 = ts.get_tick_data('000852', date='2018-04-09')
z = df.type == '买盘'
df.time[1]

df.sort_values('time', inplace=True)
df.groupby('type').mean()
t = df[df.volume > 400]

junzhi = (df.price * df.volume).sum() / df.volume.sum()

t1 = df[(df.type == '买盘') & (df.volume > 400)]
junzhi1 = (t1.price * t1.volume).sum() / t1.volume.sum()

import time

z = (time.strptime('09:31:21', '%H:%M:%S'))
df.head(10)

df.time.to_datetime()
import pandas as pd

df.time = pd.to_datetime(df.time, format='%H:%M:%S').time()
df.time = [x.time() for x in df.time]
df.set_index('time', inplace=True)
df.time[1]
df = ts.get_realtime_quotes('000581')

df2 = ts.get_sina_dd('000852', date='2018-04-09')

import os


def file_name(key, file_dir='C:\\Users\\chenzq\\Documents\\9_23'):
    temp = []
    for root, dirs, files in os.walk(file_dir):
        for file in files:
            if os.path.splitext(file)[1] in ['.py', '.txt']:
                temp.append(os.path.join(root, file))
    for dir in temp:
        s = 0
        # dir=file[8]
        try:
            with open(dir, encoding='utf-8') as f:
                for i in f:
                    if key in i:
                        print(i),
                        s += 1
                # if s == 0:
                # print("don't match it！")
        except:
            try:
                with open(dir, encoding='gbk') as f:
                    for i in f:
                        if key in i:
                            print(i),
                            s += 1
                    # if s == 0:
                    # print("don't match it！")
            except:
                print(dir)


file_name('apply')

#################################导出到xlsx工作簿sheet
import xlrd

with xlrd.get_excel('temp1.xlsx') as e:
    temp1.to_excel(e, 'D0')

from xlutils.copy import copy
from xlrd import open_workbook

w = copy(xlrd.open_workbook('temp1.xls'))
w.get_sheet('dd').write(0, 0, "foo")
w.save('book2.xls')

from openpyxl import load_workbook

book = load_workbook('temp1.xlsx')
writer = pd.ExcelWriter('temp1.xlsx', engine='openpyxl')
writer.book = book
temp1.to_excel('temp1.xlsx', 'sheetf11')
writer.save()

temp1.to_excel(writer, 'sheet111')
pd.DataFrame(userProfile, index=[1]).to_excel(writer, 'sheet123', startrow=0, startcol=0)
writer.save()
pd.DataFrame(userProfile, index=[1]).to_excel(writer, 'sheet123', startrow=3, startcol=3)
writer.save()

data_filtered.to_excel(writer, "Main", cols=['Diff1', 'Diff2'])

writer.save()
temp1.to_excel('tem.xls', 'Sheet43')

from openpyxl import load_workbook

book = xlrd.open_workbook('temp1.xls')
writer = pd.ExcelWriter('temp1.xls', engine='openpyxl')
writer.book = book
temp1.to_excel('temp1.xlsx', 'sheet111')
writer.save()
